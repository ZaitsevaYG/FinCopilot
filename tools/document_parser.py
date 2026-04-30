from langchain.tools import tool
import openpyxl
import pandas as pd
from pathlib import Path
import pdfplumber
import json

@tool
def parse_inputs(query: str) -> str:
    """Парсит Excel/PDF из inputs/, извлекает таблицы plan/fact/справочники"""
    data = {}
    inputs_path = Path("inputs")
    if not inputs_path.exists():
        return json.dumps({"error": "Нет файлов в inputs/"})
    
    for file in inputs_path.glob("*"):
        if file.suffix == ".xlsx":
            try:
                df = pd.read_excel(file, sheet_name=None)
                data[str(file)] = {
                    "sheets": {sheet: df[sheet].to_dict('records')[:10] for sheet in df},  # первые 10 строк
                    "summary": f"Листы: {list(df.keys())}, строки: {sum(len(s) for s in df.values())}"
                }
            except:
                wb = openpyxl.load_workbook(file)
                data[str(file)] = f"Листы: {wb.sheetnames}, активный: {sum(cell.value or 0 for row in wb.active.iter_rows(values_only=True))}"
        elif file.suffix.lower() in ['.pdf']:
            with pdfplumber.open(file) as pdf:
                text = pdf.pages[0].extract_text()[:1000]
                tables = [page.extract_table() for page in pdf.pages[:3]]
                data[str(file)] = {"text": text, "tables": tables}
    return json.dumps(data, ensure_ascii=False)