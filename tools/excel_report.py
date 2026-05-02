import json
import re
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =========================
# Config
# =========================
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


# =========================
# Helpers
# =========================
def _to_obj(raw_inputs: Any) -> Dict[str, Any]:
    if isinstance(raw_inputs, dict):
        return raw_inputs
    if isinstance(raw_inputs, str):
        text = raw_inputs.strip()
        if not text:
            return {}
        try:
            return json.loads(text)
        except Exception:
            return {"raw_text": text}
    return {"raw_text": str(raw_inputs)}


def _safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]\*\:/\\\?]", "_", str(name))
    name = name.strip() or "sheet"
    return name[:31]


def _find_metric(data: Any, target_key: str):
    if isinstance(data, dict):
        if target_key in data and data[target_key] not in [None, ""]:
            return data[target_key]
        for value in data.values():
            found = _find_metric(value, target_key)
            if found not in [None, ""]:
                return found
    elif isinstance(data, list):
        for item in data:
            found = _find_metric(item, target_key)
            if found not in [None, ""]:
                return found
    return None


def _as_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(" ", "").replace(",", ".")
    m = re.search(r"-?\d+(\.\d+)?", text)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _first_non_null(*values):
    for v in values:
        if v not in [None, "", [], {}]:
            return v
    return None


def _extract_metrics(data: Dict[str, Any]) -> Dict[str, Any]:
    metrics_root = data.get("metrics", {}) if isinstance(data, dict) else {}

    result = {
        "project_name": _first_non_null(
            metrics_root.get("project_name"),
            _find_metric(data, "project_name"),
        ),
        "budget_plan": _first_non_null(
            metrics_root.get("budget_plan"),
            _find_metric(data, "budget_plan"),
        ),
        "budget_fact": _first_non_null(
            metrics_root.get("budget_fact"),
            _find_metric(data, "budget_fact"),
        ),
        "volume_plan": _first_non_null(
            metrics_root.get("volume_plan"),
            _find_metric(data, "volume_plan"),
        ),
        "volume_fact": _first_non_null(
            metrics_root.get("volume_fact"),
            _find_metric(data, "volume_fact"),
        ),
        "payback_plan": _first_non_null(
            metrics_root.get("payback_plan"),
            _find_metric(data, "payback_plan"),
        ),
        "payback_fact": _first_non_null(
            metrics_root.get("payback_fact"),
            _find_metric(data, "payback_fact"),
        ),
    }

    for key in ["budget_plan", "budget_fact", "volume_plan", "volume_fact", "payback_plan", "payback_fact"]:
        result[key] = _as_float(result[key])

    return result


def _compute_summary(metrics: Dict[str, Any]) -> Dict[str, Any]:
    budget_plan = metrics.get("budget_plan")
    budget_fact = metrics.get("budget_fact")
    volume_plan = metrics.get("volume_plan")
    volume_fact = metrics.get("volume_fact")
    payback_plan = metrics.get("payback_plan")
    payback_fact = metrics.get("payback_fact")

    budget_delta = (budget_fact - budget_plan) if budget_plan is not None and budget_fact is not None else None
    budget_delta_pct = ((budget_fact - budget_plan) / budget_plan * 100) if budget_plan not in [None, 0] and budget_fact is not None else None

    volume_delta = (volume_fact - volume_plan) if volume_plan is not None and volume_fact is not None else None
    volume_delta_pct = ((volume_fact - volume_plan) / volume_plan * 100) if volume_plan not in [None, 0] and volume_fact is not None else None

    payback_delta = (payback_fact - payback_plan) if payback_plan is not None and payback_fact is not None else None

    return {
        "budget_delta": budget_delta,
        "budget_delta_pct": budget_delta_pct,
        "volume_delta": volume_delta,
        "volume_delta_pct": volume_delta_pct,
        "payback_delta": payback_delta,
    }


def _build_summary_df(query: str, metrics: Dict[str, Any], summary: Dict[str, Any], plan: Dict[str, Any]) -> pd.DataFrame:
    rows = [
        ["Дата формирования", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Запрос пользователя", query],
        ["Тип анализа", plan.get("intent", "n/a")],
        ["Полный отчет", bool(plan.get("full_report", False))],
        ["Нужен текст доклада", bool(plan.get("need_report_text", False))],
        ["Проект", metrics.get("project_name") or "н/д"],
        ["Бюджет план", metrics.get("budget_plan")],
        ["Бюджет факт", metrics.get("budget_fact")],
        ["Отклонение бюджета", summary.get("budget_delta")],
        ["Отклонение бюджета, %", summary.get("budget_delta_pct")],
        ["Объем план", metrics.get("volume_plan")],
        ["Объем факт", metrics.get("volume_fact")],
        ["Отклонение объема", summary.get("volume_delta")],
        ["Отклонение объема, %", summary.get("volume_delta_pct")],
        ["Окупаемость план", metrics.get("payback_plan")],
        ["Окупаемость факт", metrics.get("payback_fact")],
        ["Отклонение окупаемости", summary.get("payback_delta")],
    ]
    return pd.DataFrame(rows, columns=["Показатель", "Значение"])


def _build_plan_fact_df(metrics: Dict[str, Any], summary: Dict[str, Any]) -> pd.DataFrame:
    rows = [
        ["Бюджет", metrics.get("budget_plan"), metrics.get("budget_fact"), summary.get("budget_delta"), summary.get("budget_delta_pct")],
        ["Объем", metrics.get("volume_plan"), metrics.get("volume_fact"), summary.get("volume_delta"), summary.get("volume_delta_pct")],
        ["Окупаемость", metrics.get("payback_plan"), metrics.get("payback_fact"), summary.get("payback_delta"), None],
    ]
    return pd.DataFrame(rows, columns=["Показатель", "План", "Факт", "Отклонение", "Отклонение_%"])


def _build_factor_analysis_df(metrics: Dict[str, Any], summary: Dict[str, Any], plan: Dict[str, Any]) -> pd.DataFrame:
    rows = []

    if summary.get("budget_delta") is not None:
        rows.append(["Бюджет", "Финансовое отклонение по бюджету", summary.get("budget_delta"), summary.get("budget_delta_pct")])

    if summary.get("volume_delta") is not None:
        rows.append(["Объем", "Отклонение по объему производства/выпуска", summary.get("volume_delta"), summary.get("volume_delta_pct")])

    if summary.get("payback_delta") is not None:
        rows.append(["Окупаемость", "Изменение срока окупаемости", summary.get("payback_delta"), None])

    if not rows:
        rows.append(["Нет данных", "Недостаточно данных для факторного анализа", None, None])

    return pd.DataFrame(rows, columns=["Фактор", "Описание", "Влияние", "Влияние_%"])


def _build_liquidity_df(metrics: Dict[str, Any], summary: Dict[str, Any]) -> pd.DataFrame:
    budget_plan = metrics.get("budget_plan")
    budget_fact = metrics.get("budget_fact")
    volume_fact = metrics.get("volume_fact")

    coverage_ratio = None
    cost_per_unit_fact = None

    if budget_plan not in [None, 0] and budget_fact is not None:
        coverage_ratio = budget_fact / budget_plan

    if volume_fact not in [None, 0] and budget_fact is not None:
        cost_per_unit_fact = budget_fact / volume_fact

    rows = [
        ["Коэффициент покрытия бюджета (факт/план)", coverage_ratio],
        ["Условная капиталоемкость факта (бюджет факт / объем факт)", cost_per_unit_fact],
        ["Комментарий", "Лист расчетный, показатели строятся только из доступных входных данных"],
    ]
    return pd.DataFrame(rows, columns=["Показатель", "Значение"])


def _build_costing_df(metrics: Dict[str, Any]) -> pd.DataFrame:
    cost_plan = None
    cost_fact = None

    if metrics.get("budget_plan") not in [None, 0] and metrics.get("volume_plan") not in [None, 0]:
        cost_plan = metrics["budget_plan"] / metrics["volume_plan"]

    if metrics.get("budget_fact") not in [None, 0] and metrics.get("volume_fact") not in [None, 0]:
        cost_fact = metrics["budget_fact"] / metrics["volume_fact"]

    delta = None
    delta_pct = None

    if cost_plan is not None and cost_fact is not None:
        delta = cost_fact - cost_plan
        if cost_plan != 0:
            delta_pct = delta / cost_plan * 100

    rows = [
        ["Себестоимость/ед. план", cost_plan],
        ["Себестоимость/ед. факт", cost_fact],
        ["Отклонение", delta],
        ["Отклонение_%", delta_pct],
    ]
    return pd.DataFrame(rows, columns=["Показатель", "Значение"])


def _build_waterfall_df(metrics: Dict[str, Any], summary: Dict[str, Any], plan: Dict[str, Any]) -> pd.DataFrame:
    rows = [
        ["Плановый бюджет", metrics.get("budget_plan") or 0],
        ["Влияние объема", summary.get("volume_delta") or 0],
        ["Прочие факторы", summary.get("budget_delta") or 0],
        ["Фактический бюджет", metrics.get("budget_fact") or 0],
    ]
    return pd.DataFrame(rows, columns=["Этап", "Значение"])


def _build_raw_metrics_df(data: Dict[str, Any]) -> pd.DataFrame:
    metrics = data.get("metrics", {}) if isinstance(data, dict) else {}
    if not metrics:
        return pd.DataFrame([{"Показатель": "metrics", "Значение": "нет данных"}])

    rows = [{"Показатель": k, "Значение": v} for k, v in metrics.items()]
    return pd.DataFrame(rows)


def _build_files_df(data: Dict[str, Any]) -> pd.DataFrame:
    files = data.get("files", []) if isinstance(data, dict) else []
    rows = []

    for f in files:
        rows.append({
            "file_name": f.get("file_name"),
            "file_type": f.get("file_type"),
            "error": f.get("error"),
            "warning": f.get("warning"),
        })

    if not rows:
        rows = [{"file_name": "n/a", "file_type": "n/a", "error": None, "warning": "Нет данных о файлах"}]

    return pd.DataFrame(rows)


def _build_numeric_hints_df(data: Dict[str, Any]) -> pd.DataFrame:
    hints = data.get("numeric_hints", []) if isinstance(data, dict) else []
    if not hints:
        return pd.DataFrame([{"path": "n/a", "value": None}])
    return pd.DataFrame(hints)


# =========================
# Workbook formatting
# =========================
def _format_worksheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_length = 0
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)

        for cell in col:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 14), 40)


def _apply_number_formats(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'


def _add_waterfall_chart(ws):
    max_row = ws.max_row
    if max_row < 2:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Waterfall Data"
    chart.y_axis.title = "Значение"
    chart.x_axis.title = "Этап"

    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width 